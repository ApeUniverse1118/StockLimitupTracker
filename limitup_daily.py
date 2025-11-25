import akshare as ak
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
import os

"""
=========================================================
文件名称: limitup_tracker.py
功能说明: 获取每日股票涨停数据，统计关键指标，并写入 Excel 文件
依赖库: akshare, pandas, openpyxl
作者: Jason Zhao
日期: 2025-11-25
版本: 1.0
更新日志:
  - 1.0: 完成基础功能，包括获取涨停数据、统计每日指标、写入 Excel
使用说明:
  1. 安装依赖库:
       pip install akshare pandas openpyxl
  2. 运行脚本:
       python limitup_tracker.py
  3. 输出文件:
       - 导出目录: ./limitup_tracker/
       - 文件名格式: limitup_tracker_YYYYMMDD.xlsx
       - Sheet:
           * LimitUp_Detail: 涨停明细
           * Daily_Summary: 每日汇总统计
=========================================================
"""

# ----------------------------
#  配置
# ----------------------------
FILE_BASE = "limitup_tracker"  # 文件名基础
EXPORT_DIR = "limitup_tracker"   # 导出根目录
DETAIL_SHEET = "LimitUp_Detail"
SUMMARY_SHEET = "Daily_Summary"

# ----------------------------
# 获取上一交易日日期
# ----------------------------
def get_last_trade_date():
    # 简单方法：昨天
    yesterday = datetime.now() - timedelta(days=1)
    return yesterday.strftime("%Y%m%d")

# ----------------------------
# 构建完整文件路径
# ----------------------------
def get_file_path():
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)  # 创建目录
    last_date = get_last_trade_date()
    file_name = f"{FILE_BASE}_{last_date}.xlsx"
    full_path = os.path.join(EXPORT_DIR, file_name)
    return full_path

# ----------------------------
# 获取当日涨停数据
# ----------------------------
def get_limit_up_data(dateStr):
    try:
        df = ak.stock_zt_pool_em(dateStr)  # 东方财富涨停池
        print("获取到涨停数据，条数:", len(df))
        print("样例数据:\n", df.head())
        df["流通市值(亿)"] = df["流通市值"].astype(float) / 1e8
        df["日期"] = datetime.now().strftime("%Y-%m-%d")
        return df
    except Exception as e:
        print("获取涨停数据失败:", e)
        return pd.DataFrame()

# ----------------------------
# 统计每日指标
# ----------------------------
def compute_daily_summary(df):
    if df.empty:
        return None

    date = df["日期"].iloc[0]
    total = len(df)

    if "连板数" in df.columns:
        df["连板数"] = df["连板数"].fillna(1)
    else:
        df["连板数"] = 1

    max_board = df["连板数"].max()
    lianban_count = (df["连板数"] > 1).sum()
    first_board = (df["连板数"] == 1).sum()

    top_industry = df["所属行业"].value_counts().head(3)
    industries = list(top_industry.index)
    counts = list(top_industry.values)

    avg_mv = df["流通市值(亿)"].mean()
    med_mv = df["流通市值(亿)"].median()

    summary = {
        "日期": date,
        "涨停家数": total,
        "连板家数": lianban_count,
        "最高板高度": max_board,
        "首板数量": first_board,
        "平均流通市值(亿)": round(avg_mv, 2),
        "中位流通市值(亿)": round(med_mv, 2),
        "主线板块1": industries[0] if len(industries) > 0 else "",
        "数量1": counts[0] if len(counts) > 0 else "",
        "主线板块2": industries[1] if len(industries) > 1 else "",
        "数量2": counts[1] if len(counts) > 1 else "",
        "主线板块3": industries[2] if len(industries) > 2 else "",
        "数量3": counts[2] if len(counts) > 2 else "",
    }

    return pd.DataFrame([summary])

# ----------------------------
# 向 Excel 写入数据
# ----------------------------
def write_to_excel(df_detail, df_summary, file_path):
    if not os.path.exists(file_path):
        # 文件不存在，直接用 pandas 写入，自动创建 sheet
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_detail.to_excel(writer, sheet_name=DETAIL_SHEET, index=False)
            df_summary.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)
    else:
        # 文件存在，追加写入
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df_detail.to_excel(
                writer,
                sheet_name=DETAIL_SHEET,
                index=False,
                header=writer.sheets[DETAIL_SHEET].max_row == 1,
                startrow=writer.sheets[DETAIL_SHEET].max_row
            )
            df_summary.to_excel(
                writer,
                sheet_name=SUMMARY_SHEET,
                index=False,
                header=writer.sheets[SUMMARY_SHEET].max_row == 1,
                startrow=writer.sheets[SUMMARY_SHEET].max_row
            )

# ----------------------------
# 主流程
# ----------------------------
def main():
    last_date = get_last_trade_date()
    file_path = get_file_path()
    print("获取上一交易日数据:", last_date)
    df = get_limit_up_data(dateStr=last_date)

    if df.empty:
        print("今日无涨停数据或获取失败。")
        return

    print("计算统计指标...")
    df_summary = compute_daily_summary(df)

    print("写入 Excel...")
    write_to_excel(df, df_summary, file_path)

    print("任务完成！已写入:", file_path)


if __name__ == "__main__":
    main()
